import openpyxl, json, sys, datetime

import os
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser("~/Downloads/Stock PP 2025.xlsx")
OUT = os.path.join(os.path.dirname(__file__), ".import-data.json")
wb = openpyxl.load_workbook(SRC, data_only=True)

TYPES = {"buy": "buy", "sell": "sell", "sample": "sample", "scrap": "scrap"}

def numstr(v):
    if v is None or v == "":
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", "")
        if v == "" or v == "-":
            return None
    try:
        d = float(v)
    except Exception:
        return None
    d = round(d, 6)  # kill float noise (e.g. 19.100000000000364 -> 19.1)
    if d == int(d):
        return str(int(d))
    return ("%.6f" % d).rstrip("0").rstrip(".")

def datestr(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return None

def split_brand_grade(tab):
    parts = tab.strip().split(" ", 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return "", tab.strip()

out = []
report = []
for name in wb.sheetnames:
    if name.strip().lower() == "summary":
        continue
    ws = wb[name]
    # find header row containing 'Tanggal'
    hdr = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        if any(isinstance(c, str) and c.strip().lower() == "tanggal" for c in row):
            hdr = i
            break
    if hdr is None:
        report.append((name, "NO_HEADER", 0))
        continue
    data_start = hdr + 2  # skip Tanggal row + Unit/Harga subrow
    brand, grade = split_brand_grade(name)
    raw = []
    skipped = 0
    for row in ws.iter_rows(min_row=data_start, min_col=2, max_col=14, values_only=True):
        # offsets: 0 B Tanggal,1 C docNo,2 D type,3 E cust,4 F PembUnit,5 G PembHarga,
        # 6 H PembJumlah,7 I PenjUnit,8 J PenjHarga,9 K PenjJumlah,10 L SaldoUnit
        if row is None:
            continue
        pemb_unit = numstr(row[4])
        penj_unit = numstr(row[7])
        rtype = row[2]
        t = TYPES.get(rtype.strip().lower()) if isinstance(rtype, str) else None
        if t is None:
            # blank/unknown type: infer from which side has a quantity (opening / continuation rows)
            if pemb_unit and pemb_unit != "0":
                t = "buy"
            elif penj_unit and penj_unit != "0":
                t = "sell"
            else:
                skipped += 1
                continue
        if t == "buy":
            # some sheets mis-log buys in the Penjualan columns (type says Buy, values on sell side)
            qty = pemb_unit or penj_unit
            unit = numstr(row[5]) or numstr(row[8])
        else:
            qty = penj_unit or pemb_unit
            unit = None
        if qty is None or qty == "0":
            skipped += 1
            continue
        raw.append({
            "date": datestr(row[0]),  # may be None
            "type": t, "qty": qty, "unitCost": unit, "salePrice": None,
            "docNo": (str(row[1]).strip() if row[1] not in (None, "") else None),
            "counterparty": (str(row[3]).strip() if row[3] not in (None, "") else None),
        })

    # forward-fill missing dates; back-fill leading None with first known date
    first_known = next((r["date"] for r in raw if r["date"]), None)
    last = first_known
    for r in raw:
        r["date"] = r["date"] or last
        last = r["date"]
    # clamp monotonically non-decreasing to preserve sheet (chronological) order
    mx = None
    for r in raw:
        if r["date"] is None:
            continue
        if mx and r["date"] < mx:
            r["date"] = mx
        else:
            mx = r["date"]
    txns = [r for r in raw if r["date"]]

    if txns:
        out.append({"brand": brand, "grade": grade, "txns": txns})
        report.append((name, "OK", len(txns), skipped))
    else:
        report.append((name, "EMPTY", 0, skipped))

with open(OUT, "w") as f:
    json.dump(out, f)

print(f"sheets with data: {len(out)}")
total = sum(len(m['txns']) for m in out)
print(f"total transactions: {total}")
print("\n--- per sheet ---")
for r in report:
    print(r)
